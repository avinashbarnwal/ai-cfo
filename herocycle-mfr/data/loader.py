"""Data access layer for HeroCycle MFR - Excel edition.

Input is the two normal month-end close files:
  - HeroCycle_Financials_Jun2026.xlsx  (P&L Summary, Monthly Trend, EBITDA Bridge, Cash & WC)
  - HeroCycle_Drivers_Jun2026.xlsx     (Revenue by Channel, Revenue by Segment, Operating KPIs)

This module is the ONLY place that knows workbook structure. It infers
sheets by fuzzy name match and rows by label match, so reasonable
variations in file layout still parse. The rest of the system (agent,
API, UI) consumes the canonical dataset dict returned by load_dataset().

Swap the workbook paths each month; nothing else changes.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent
FINANCIALS_FILE = DATA_DIR / "HeroCycle_Financials_Jun2026.xlsx"
DRIVERS_FILE = DATA_DIR / "HeroCycle_Drivers_Jun2026.xlsx"


# ------------------------------------------------------------------ helpers
def _find_sheet(wb, *keywords: str):
    """Return the first sheet whose name contains all keywords (case-insensitive)."""
    for name in wb.sheetnames:
        low = name.lower()
        if all(k.lower() in low for k in keywords):
            return wb[name]
    return None


def _header_row(ws, first_header: str) -> int | None:
    """Find the header row: first row whose first cell EQUALS first_header
    (case-insensitive). Exact match avoids catching title rows whose text
    merely contains the word (e.g. 'Revenue by Channel - June 2026')."""
    for row in ws.iter_rows(min_row=1, max_row=10):
        v = row[0].value
        if isinstance(v, str) and v.strip().lower() == first_header.lower():
            return row[0].row
    return None


def _rows_as_dicts(ws, header_row: int) -> list[dict[str, Any]]:
    """Read the table under header_row until the first fully empty row."""
    headers = [c.value for c in ws[header_row] if c.value is not None]
    out = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=len(headers)):
        values = [c.value for c in row]
        if all(v is None for v in values):
            break
        out.append(dict(zip(headers, values)))
    return out


def _label_value_pairs(ws, header_row: int) -> list[tuple[str, float]]:
    pairs = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=2):
        label, value = row[0].value, row[1].value
        if label is None and value is None:
            break
        if isinstance(label, str) and isinstance(value, (int, float)):
            pairs.append((label.strip(), round(float(value), 2)))
    return pairs


def _find(rows: list[dict], key_col: str, label: str) -> dict | None:
    for r in rows:
        v = r.get(key_col)
        if isinstance(v, str) and label.lower() in v.lower():
            return r
    return None


# ------------------------------------------------------------------ parsers
def _parse_financials(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)  # cached values; workbooks must be recalculated
    out: dict[str, Any] = {}

    # P&L Summary
    ws = _find_sheet(wb, "p&l") or _find_sheet(wb, "pnl") or _find_sheet(wb, "summary")
    if ws is not None:
        hr = _header_row(ws, "Line item")
        rows = _rows_as_dicts(ws, hr) if hr else []

        def pnl(label, scale=1.0):
            r = _find(rows, "Line item", label)
            if not r:
                return None
            return {
                "actual": round(r.get("Actual", 0) * scale, 2),
                "budget": round(r.get("Budget", 0) * scale, 2),
                "prior_month": round(r.get("Prior Month", 0) * scale, 2),
                "prior_year": round(r.get("Prior Year", 0) * scale, 2),
            }

        out["pnl"] = {
            "revenue": pnl("Revenue"),
            "gross_margin_pct": pnl("Gross margin", scale=100),  # stored as fraction in xlsx
            "opex": pnl("Opex"),
            "ebitda": pnl("EBITDA"),
        }

    # Monthly Trend
    ws = _find_sheet(wb, "trend")
    if ws is not None:
        hr = _header_row(ws, "Month")
        out["trend"] = [
            {
                "month": r["Month"],
                "revenue": round(r["Revenue"], 2),
                "budget": round(r["Budget"], 2),
                "gm_pct": round(r["GM %"] * 100, 1),
            }
            for r in (_rows_as_dicts(ws, hr) if hr else [])
        ]

    # EBITDA Bridge
    ws = _find_sheet(wb, "ebitda", "bridge") or _find_sheet(wb, "bridge")
    if ws is not None:
        hr = _header_row(ws, "Bridge item")
        pairs = _label_value_pairs(ws, hr) if hr else []
        out["ebitda_bridge"] = [
            {"label": lab, "value": val, "is_total": "ebitda" in lab.lower()}
            for lab, val in pairs
        ]

    # Cash & WC
    ws = _find_sheet(wb, "cash")
    if ws is not None:
        hr = _header_row(ws, "Item")
        pairs = _label_value_pairs(ws, hr) if hr else []
        opening = next((v for l, v in pairs if "opening" in l.lower()), None)
        closing = next((v for l, v in pairs if "closing" in l.lower()), None)
        bridge = [
            {"label": l, "value": v}
            for l, v in pairs
            if "opening" not in l.lower() and "closing" not in l.lower()
        ]
        out["cash"] = {"opening": opening, "closing": closing, "bridge": bridge}

    return out


def _parse_drivers(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    out: dict[str, Any] = {}

    ws = _find_sheet(wb, "channel")
    if ws is not None:
        hr = _header_row(ws, "Channel")
        out["channels"] = [
            {
                "name": r["Channel"],
                "actual": round(r["Actual"], 2),
                "budget": round(r["Budget"], 2),
                "gm_pct": round(r["GM %"] * 100, 1),
            }
            for r in (_rows_as_dicts(ws, hr) if hr else [])
            if isinstance(r.get("Channel"), str) and r["Channel"].lower() != "total"
        ]

    ws = _find_sheet(wb, "segment")
    if ws is not None:
        hr = _header_row(ws, "Segment")
        out["segments"] = [
            {
                "name": r["Segment"],
                "actual": round(r["Actual"], 2),
                "budget": round(r["Budget"], 2),
                "growth_yoy_pct": round(r["YoY growth %"] * 100, 1),
            }
            for r in (_rows_as_dicts(ws, hr) if hr else [])
            if isinstance(r.get("Segment"), str) and r["Segment"].lower() != "total"
        ]

    ws = _find_sheet(wb, "kpi")
    if ws is not None:
        hr = _header_row(ws, "KPI")
        rows = _rows_as_dicts(ws, hr) if hr else []
        units = _find(rows, "KPI", "Units")
        aov = _find(rows, "KPI", "order value")
        mkt = _find(rows, "KPI", "Marketing")
        out["drivers"] = {
            "units": units and int(units["Actual"]),
            "units_budget": units and int(units["Budget"]),
            "aov_inr": aov and int(aov["Actual"]),
            "aov_budget_inr": aov and int(aov["Budget"]),
            "marketing_spend": mkt and round(mkt["Actual"], 2),
            "marketing_budget": mkt and round(mkt["Budget"], 2),
        }

    return out


# ------------------------------------------------------------------ public API
def load_dataset(
    financials_path: Path | str | None = None,
    drivers_path: Path | str | None = None,
) -> dict[str, Any]:
    """Parse the two month-end workbooks into the canonical dataset dict."""
    fin = Path(financials_path or FINANCIALS_FILE)
    drv = Path(drivers_path or DRIVERS_FILE)
    data: dict[str, Any] = {
        "company": "HeroCycle",
        "period": "June 2026 (FY27 Q1 close)",
        "currency": "INR Cr",
        "source_files": {"financials": fin.name, "drivers": drv.name},
    }
    data.update(_parse_financials(fin))
    data.update(_parse_drivers(drv))
    return data


def validate_dataset(data: dict[str, Any]) -> dict[str, Any]:
    """S1 input-validation checkpoint: coverage, comparisons, gaps."""
    required = ["pnl", "channels", "segments", "drivers", "cash", "trend"]
    missing = [k for k in required if not data.get(k)]
    comparisons = []
    rev = (data.get("pnl") or {}).get("revenue") or {}
    if "budget" in rev:
        comparisons.append("Budget")
    if "prior_month" in rev:
        comparisons.append("Prior month")
    if "prior_year" in rev:
        comparisons.append("Prior year")
    if data.get("trend"):
        comparisons.append(f"{len(data['trend'])}-month trend")
    return {
        "company": data.get("company"),
        "period": data.get("period"),
        "currency": data.get("currency"),
        "source_files": data.get("source_files"),
        "sections_found": [k for k in required if data.get(k)],
        "usable_comparisons": comparisons,
        "data_gaps": ["Balance sheet detail", "Headcount by function"],
        "missing_critical": missing,
        "can_proceed": not missing,
    }


def chart_payload(chart_key: str, data: dict[str, Any]) -> dict[str, Any]:
    """Render-ready chart data for the UI. The UI only knows how to
    draw a payload of a given kind; it never touches workbook structure."""
    if chart_key == "rev_trend":
        return {"kind": "trend", "series": data["trend"]}
    if chart_key == "channel_bar":
        return {"kind": "grouped_bar", "rows": data["channels"]}
    if chart_key == "ebitda_bridge":
        return {"kind": "bridge", "rows": data["ebitda_bridge"]}
    if chart_key == "cash_bridge":
        rows = (
            [{"label": "Opening cash", "value": data["cash"]["opening"], "is_total": True}]
            + data["cash"]["bridge"]
            + [{"label": "Closing cash", "value": data["cash"]["closing"], "is_total": True}]
        )
        return {"kind": "bridge", "rows": rows}
    raise KeyError(f"Unknown chart_key: {chart_key}")


if __name__ == "__main__":
    import json

    ds = load_dataset()
    print(json.dumps(validate_dataset(ds), indent=2))
    print(json.dumps(ds, indent=2)[:1200])
