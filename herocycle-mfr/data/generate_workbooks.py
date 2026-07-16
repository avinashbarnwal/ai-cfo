"""Generate HeroCycle month-end workbooks (June 2026).

Produces the two normal finance files the MFR expects:
  1. HeroCycle_Financials_Jun2026.xlsx  (P&L, trend, EBITDA bridge, cash & WC)
  2. HeroCycle_Drivers_Jun2026.xlsx     (channels, segments, operating KPIs)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1D2025")
HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=ARIAL, bold=True, size=12)
BASE = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
BLUE = Font(name=ARIAL, size=10, color="0000FF")  # hardcoded inputs
NOTE = Font(name=ARIAL, size=9, italic=True, color="6B7078")
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))

CR = '#,##0.0;(#,##0.0);"-"'
PCT = "0.0%"
INT = '#,##0;(#,##0);"-"'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


# ---------------------------------------------------------------- financials
wb = Workbook()

# --- P&L Summary ---
ws = wb.active
ws.title = "P&L Summary"
ws["A1"] = "HeroCycle - P&L Summary - June 2026 (INR Cr)"
ws["A1"].font = TITLE_FONT
headers = ["Line item", "Actual", "Budget", "Prior Month", "Prior Year", "Var vs Budget", "Var vs Budget %"]
for c, h in enumerate(headers, start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(headers))

pnl_rows = [
    ("Revenue", 142.6, 135.0, 138.2, 121.4, CR),
    ("Gross margin %", 0.278, 0.295, 0.291, 0.286, PCT),
    ("Opex", 28.4, 27.5, 27.1, 25.9, CR),
    ("EBITDA", 11.2, 12.4, 13.1, 9.8, CR),
]
r = 4
for name, act, bud, pm, py, fmt in pnl_rows:
    ws.cell(row=r, column=1, value=name).font = BOLD
    for c, v in zip(range(2, 6), [act, bud, pm, py]):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = BLUE
        cell.number_format = fmt
    ws.cell(row=r, column=6, value=f"=B{r}-C{r}").number_format = fmt
    vc = ws.cell(row=r, column=7, value=f"=IFERROR((B{r}-C{r})/C{r},0)")
    vc.number_format = PCT
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = THIN
        if ws.cell(row=r, column=c).font not in (BOLD, BLUE):
            ws.cell(row=r, column=c).font = BASE
    r += 1
ws["A9"] = "Blue cells: month-end close inputs. Black cells: formulas. Source: HeroCycle finance close pack, June 2026 (demo data)."
ws["A9"].font = NOTE
set_widths(ws, [22, 12, 12, 13, 12, 14, 16])

# --- Monthly Trend ---
ws = wb.create_sheet("Monthly Trend")
ws["A1"] = "Monthly Trend - FY26 H2 to FY27 Q1 (INR Cr)"
ws["A1"].font = TITLE_FONT
for c, h in enumerate(["Month", "Revenue", "Budget", "GM %"], start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 4)
trend = [
    ("Jan", 118.4, 120.0, 0.294), ("Feb", 121.9, 122.5, 0.292), ("Mar", 129.8, 126.0, 0.290),
    ("Apr", 131.5, 128.5, 0.293), ("May", 138.2, 131.5, 0.291), ("Jun", 142.6, 135.0, 0.278),
]
for i, (m, rev, bud, gm) in enumerate(trend, start=4):
    ws.cell(row=i, column=1, value=m).font = BASE
    ws.cell(row=i, column=2, value=rev).font = BLUE
    ws.cell(row=i, column=2).number_format = CR
    ws.cell(row=i, column=3, value=bud).font = BLUE
    ws.cell(row=i, column=3).number_format = CR
    ws.cell(row=i, column=4, value=gm).font = BLUE
    ws.cell(row=i, column=4).number_format = PCT
set_widths(ws, [10, 12, 12, 10])

# --- EBITDA Bridge ---
ws = wb.create_sheet("EBITDA Bridge")
ws["A1"] = "EBITDA Bridge - Budget to Actual - June 2026 (INR Cr)"
ws["A1"].font = TITLE_FONT
for c, h in enumerate(["Bridge item", "Impact"], start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 2)
bridge = [
    ("Budget EBITDA", 12.4), ("Revenue volume beat", 2.1), ("Channel mix (marketplace/D2C)", -1.6),
    ("Input cost & discounting", -0.8), ("Marketing timing", -1.2), ("Other opex", 0.3),
]
for i, (label, v) in enumerate(bridge, start=4):
    ws.cell(row=i, column=1, value=label).font = BASE
    cell = ws.cell(row=i, column=2, value=v)
    cell.font = BLUE
    cell.number_format = CR
ws.cell(row=10, column=1, value="Actual EBITDA").font = BOLD
tc = ws.cell(row=10, column=2, value="=SUM(B4:B9)")
tc.font = BOLD
tc.number_format = CR
ws["A12"] = "Bridge items from FP&A variance analysis of the June close (demo data)."
ws["A12"].font = NOTE
set_widths(ws, [32, 12])

# --- Cash & WC ---
ws = wb.create_sheet("Cash & WC")
ws["A1"] = "Cash & Working Capital Movement - June 2026 (INR Cr)"
ws["A1"].font = TITLE_FONT
for c, h in enumerate(["Item", "Movement"], start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 2)
ws.cell(row=4, column=1, value="Opening cash").font = BOLD
oc = ws.cell(row=4, column=2, value=41.0)
oc.font = BLUE
oc.number_format = CR
cash = [
    ("EBITDA", 11.2), ("Inventory build (pre-festive)", -5.4), ("AR increase (dealer credit)", -3.1),
    ("AP increase", 1.7), ("Capex (e-cycle line)", -8.6), ("Tax & other", -2.6),
]
for i, (label, v) in enumerate(cash, start=5):
    ws.cell(row=i, column=1, value=label).font = BASE
    cell = ws.cell(row=i, column=2, value=v)
    cell.font = BLUE
    cell.number_format = CR
ws.cell(row=11, column=1, value="Closing cash").font = BOLD
cc = ws.cell(row=11, column=2, value="=B4+SUM(B5:B10)")
cc.font = BOLD
cc.number_format = CR
set_widths(ws, [32, 12])

wb.save("HeroCycle_Financials_Jun2026.xlsx")

# ------------------------------------------------------------------ drivers
wb2 = Workbook()

ws = wb2.active
ws.title = "Revenue by Channel"
ws["A1"] = "Revenue by Channel - June 2026 (INR Cr)"
ws["A1"].font = TITLE_FONT
for c, h in enumerate(["Channel", "Actual", "Budget", "Var vs Budget", "GM %"], start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 5)
channels = [("Dealer network", 78.9, 79.5, 0.301), ("D2C online", 38.4, 32.5, 0.274), ("Marketplace", 25.3, 23.0, 0.212)]
for i, (name, act, bud, gm) in enumerate(channels, start=4):
    ws.cell(row=i, column=1, value=name).font = BASE
    ws.cell(row=i, column=2, value=act).font = BLUE
    ws.cell(row=i, column=2).number_format = CR
    ws.cell(row=i, column=3, value=bud).font = BLUE
    ws.cell(row=i, column=3).number_format = CR
    ws.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = CR
    ws.cell(row=i, column=5, value=gm).font = BLUE
    ws.cell(row=i, column=5).number_format = PCT
ws.cell(row=7, column=1, value="Total").font = BOLD
for col in ("B", "C", "D"):
    tc = ws.cell(row=7, column=ord(col) - 64, value=f"=SUM({col}4:{col}6)")
    tc.font = BOLD
    tc.number_format = CR
set_widths(ws, [20, 12, 12, 14, 10])

ws = wb2.create_sheet("Revenue by Segment")
ws["A1"] = "Revenue by Segment - June 2026 (INR Cr)"
ws["A1"].font = TITLE_FONT
for c, h in enumerate(["Segment", "Actual", "Budget", "Var vs Budget", "YoY growth %"], start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 5)
segments = [("E-cycles", 41.2, 33.8, 0.34), ("Premium", 33.8, 33.0, 0.09), ("Standard", 67.6, 68.2, -0.04)]
for i, (name, act, bud, yoy) in enumerate(segments, start=4):
    ws.cell(row=i, column=1, value=name).font = BASE
    ws.cell(row=i, column=2, value=act).font = BLUE
    ws.cell(row=i, column=2).number_format = CR
    ws.cell(row=i, column=3, value=bud).font = BLUE
    ws.cell(row=i, column=3).number_format = CR
    ws.cell(row=i, column=4, value=f"=B{i}-C{i}").number_format = CR
    ws.cell(row=i, column=5, value=yoy).font = BLUE
    ws.cell(row=i, column=5).number_format = PCT
ws.cell(row=7, column=1, value="Total").font = BOLD
for col in ("B", "C", "D"):
    tc = ws.cell(row=7, column=ord(col) - 64, value=f"=SUM({col}4:{col}6)")
    tc.font = BOLD
    tc.number_format = CR
set_widths(ws, [16, 12, 12, 14, 14])

ws = wb2.create_sheet("Operating KPIs")
ws["A1"] = "Operating KPIs - June 2026"
ws["A1"].font = TITLE_FONT
for c, h in enumerate(["KPI", "Actual", "Budget", "Unit"], start=1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 4)
kpis = [
    ("Units sold", 118400, 112000, "units", INT),
    ("Average order value", 12043, 12054, "INR", INT),
    ("Marketing spend", 6.8, 5.6, "INR Cr", CR),
]
for i, (name, act, bud, unit, fmt) in enumerate(kpis, start=4):
    ws.cell(row=i, column=1, value=name).font = BASE
    ws.cell(row=i, column=2, value=act).font = BLUE
    ws.cell(row=i, column=2).number_format = fmt
    ws.cell(row=i, column=3, value=bud).font = BLUE
    ws.cell(row=i, column=3).number_format = fmt
    ws.cell(row=i, column=4, value=unit).font = BASE
ws["A8"] = "Source: HeroCycle sales ops and marketing close reports, June 2026 (demo data)."
ws["A8"].font = NOTE
set_widths(ws, [22, 12, 12, 10])

wb2.save("HeroCycle_Drivers_Jun2026.xlsx")
print("Workbooks written.")
